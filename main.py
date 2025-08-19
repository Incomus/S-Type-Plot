import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
from datetime import datetime
import traceback
import re
from xlsxwriter import Workbook as Workbook_xw
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import numpy as np
from scipy.ndimage import gaussian_filter1d
sigma = 2

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
plt.rcParams['axes.prop_cycle'] = plt.cycler(color=['#ff7f0e',  # orange
                                                     '#2ca02c',  # green
                                                     '#1f77b4',  # blue
                                                     '#d62728',  # red
                                                     '#9467bd',  # purple
                                                     '#8c564b',  # brown
                                                     '#e377c2',  # pink
                                                     '#7f7f7f',  # gray
                                                     '#bcbd22',  # yellow-green
                                                     '#17becf']) # cyan
def collect_experiments(main_folder, depth):
    file_list = []
    
    def recursive_search(folder, current_depth, experiment_count):
        # Count Excel files in the current folder
        experiment_count += sum(1 for _ in folder.glob("*.xlsx"))
        for file in folder.glob("*.xlsx"):
            file_list.append(file)
        
        # Check if we can go deeper
        if depth == -1 or current_depth < depth:
            for subfolder in folder.iterdir():
                if subfolder.is_dir():
                    # Recursively search in the subfolder
                    experiment_count = recursive_search(subfolder, current_depth + 1, experiment_count)

        return experiment_count

    # Start the recursive search from the main folder
    experiment_count = recursive_search(main_folder, 0, 0)

    return experiment_count, file_list
    
def summarize_folders():
    while True:
        main_folder_path = input("Input path to the main folder:\n")
        main_folder = Path(main_folder_path)
        if main_folder.exists() and main_folder.is_dir():  # Check if the path exists and is a directory
            break  # Break the loop if the input is valid
        else:
            print(f"'{main_folder_path}' is not a valid directory. Please try again.")
    experiment_count, file_list = collect_experiments(main_folder, 5)
    print(f'Found {experiment_count} experiments')
    input('Press Enter to continue\n')
    return main_folder_path

def choose(text, *options):
    options = list(options)  # Convert the tuple to a list for easier handling
    if isinstance(options, list) and all(isinstance(item, list) for item in options):
        options = options[0]
    while (user_input := input(f'{text}\n({"/".join(options)})\n')) not in options:
        print('Invalid input\n')
    return user_input

def preprocess(main_folder_path, process_type):
    data, failed_files = [], []
    main_folder = Path(main_folder_path)
    experiment_count, file_list = collect_experiments(main_folder, 5)
    count = 0
    first_columns = None
    for file in file_list:
        try:
            df = pd.read_excel(file)
            file = file.stem
            if process_type == 'magnitude':
                if 'Response_Time' in file:
                    continue
            elif process_type == 'phase':
                if 'Response_Time' not in file:
                    continue
                
            file = file.replace("Response_Time_PHASE_", "") # .replace("TFall_", "TFall ").replace("TRise_", "TRise ")
            type = ''
            desc = ''
            try:
                temp = df.loc[0, 'Test Parameters:']
                type = temp[:2]
                try:
                    desc = temp[3:]
                except:
                    pass
            except:
                pass
            
            if process_type == 'phase':
                if type == '':
                    try:
                        temp = df.columns[0]
                        type = temp[:2]
                        try:
                            desc = temp[3:]
                        except:
                            pass
                    except:
                        pass
                        
                if 'Phase iteration 1' in df.columns:
                    df = df.iloc[:, 3:5]
                else:
                    df = df.iloc[:, 1:3]
            elif process_type == 'magnitude':
                df = df.iloc[:, :2]  # Keep only first two columns
            
            
            if first_columns is None:
                first_columns = df.columns.tolist()
            else:
                # Rename the columns of the current file to match the first file
                df.columns = first_columns
            
            variables = file.split('_')
            for i, var in enumerate(variables):
                df[f'variable{i+1}'] = var
            df['id'] = count
            count += 1
            
            df['ant_type'] = type
            df['comment'] = desc
            
            data.append(df)
        except Exception as e:
            print(e)
            failed_files.append((file, str(e)))

    if data:
        df = pd.concat(data, ignore_index=True)
    else:
        print("No valid data found.")
        return pd.DataFrame(), failed_files
    
    return df, failed_files

def process(df, process_type):
    df.columns = df.columns.str.lower().str.replace(" ", "_")
    if process_type == 'phase':
        cols = ['x', 'y', 'folder_name', 'ps_ant', 'exp_name', 't_type', 'voltage', 'date', 'time', 'id', 'ant_type', 'comment']
        df.columns = cols
        df.drop(columns=['folder_name', 'ps_ant', 'date', 'time'], inplace=True)
    elif process_type == 'magnitude':
        cols = ['x', 'y', 'subfolder', 'type', 'antenna', 's_param', 'measurable', 'voltage', 'date', 'time', 'id', 'ant_type', 'comment']
        df.columns = cols
        df.drop(columns=['subfolder', 'type', 's_param', 'measurable', 'date', 'time'], inplace=True)
        
        df = df.pivot(index=['antenna', 'comment', 'ant_type', 'voltage', 'id'], columns='x', values='y')
        df = df.replace({np.nan:None})
        df = df.reset_index()
    return df
    
def phase_to_excel(df):
    wb = Workbook()
    if df.empty:
        return


    df['jump_time'] = None
    df['phase_90'] = None
    df['phase_10'] = None
    df['phase_max'] = None
    
    for id in df['id'].unique():
        exp_df = df[df['id'] == id]
        exp_df = exp_df.reset_index(drop=True)
        y_max_idx = exp_df['y'].abs().idxmax()
        y_max = exp_df['y'].iloc[y_max_idx]
        
        y_10 = 0.1 * y_max
        y_90 = 0.9 * y_max
            
        row_10 = exp_df.iloc[(exp_df['y'] - y_10).abs().idxmin()]
        row_90 = exp_df.iloc[(exp_df['y'] - y_90).abs().idxmin()]
        
        x_10 = row_10['x'] #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!10=5!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        x_90 = row_90['x']
        y_10 = row_10['y']
        y_90 = row_90['y']
        x_dx = abs(x_90 - x_10)
        # Update df with the R² value
        df.loc[df['id'] == id, 'phase_max'] = y_max
        df.loc[df['id'] == id, 'phase_90'] = y_90
        df.loc[df['id'] == id, 'phase_10'] = y_10
        df.loc[df['id'] == id, 'jump_time'] = x_dx
    #df = df.drop(columns=['x', 'y']).drop_duplicates().reset_index(drop=True)
    for col in ['phase_max', 'phase_90', 'phase_10', 'jump_time']:
        df[col] = pd.to_numeric(df[col], errors='coerce').round(3)

    df = df.sort_values(by=['exp_name', 't_type', 'ant_type', 'id']).reset_index(drop=True)

    # df[['v1', 'v2']] = df['voltage'].str.extract(r'([\d\.]+)-([\d\.]+)').astype(float)
    # df['voltage'] = df.apply(lambda x: x['v2'] if x['t_type'] == 'TRise' else x['v1'], axis=1)
    # df = df.drop(columns=['v1', 'v2'])
    df = df.drop(columns=['voltage'])
    df_copy = df.copy()
    df = df.drop(columns=['x', 'y']).drop_duplicates().reset_index(drop=True)
    
    df = df.pivot(index=['exp_name', 'ant_type'], columns='t_type', values=["phase_max", "jump_time", "phase_90", "phase_10"])
    df = df.reindex(columns=["TRise", "TFall"], level=1)
    df.columns = [f"{val}_{t}" for val, t in df.columns]
    df = df.reset_index()
    
    ws = wb['Sheet']
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    start_col = len(df.columns) + 2
    start_row = 1
    unique_ids = df_copy['id'].unique()
    unique_t_types = df_copy['t_type'].unique()
    
    col_position = start_col
    for series_id in unique_ids:
        ws.cell(row=start_row, column=col_position, value=f"ID {series_id} - X")  # X column header
        ws.cell(row=start_row, column=col_position+1, value=f"ID {series_id} - Y")  # Y column header
        col_position += 2  # Move to next pair of columns
    
    max_rows = df_copy.groupby("id").size().max()  # Find max rows needed per ID
    df_copy = df_copy.sort_values(by=['x']).groupby("id")
    
    for row_index in range(max_rows):  # Iterate over the maximum row count
        col_position = start_col
        for series_id in unique_ids:
            series_data = df_copy.get_group(series_id).reset_index(drop=True)  # Data for specific ID
            if row_index < len(series_data):  # If row exists in this ID's data
                ws.cell(row=start_row + 1 + row_index, column=col_position, value=series_data.loc[row_index, 'x'])
                ws.cell(row=start_row + 1 + row_index, column=col_position+1, value=series_data.loc[row_index, 'y'])
            col_position += 2  # Move to next ID's column pair
    
    chart_loc = 2
    for t_type in unique_t_types:
        chart = ScatterChart()
        chart.title = f"{t_type} raw data"
        chart.x_axis.title = "Time, s"
        chart.y_axis.title = "Phase, deg"
        chart.x_axis.majorGridlines = ChartLines()
        chart.y_axis.majorGridlines = ChartLines()
        chart.x_axis.delete = False
        chart.y_axis.delete = False

        # Add series for each unique ID
        col_position = start_col
        for series_id in unique_ids:
            series_data = df_copy.get_group(series_id).reset_index(drop=True)  # Data for specific ID
            if series_data['t_type'].iloc[0] == t_type:
                title = f'{series_data["exp_name"].iloc[0]} {series_data["ant_type"].iloc[0]}'
                x_values = Reference(ws, min_col=col_position, min_row=2, max_row=start_row + max_rows)
                y_values = Reference(ws, min_col=col_position+1, min_row=2, max_row=start_row + max_rows)
                series = Series(y_values, x_values, title=title)
                series.smooth = False
                chart.series.append(series)
            col_position += 2  # Move to the next pair of columns

        # Position chart at J1
        chart.varyColors = False
        ws.add_chart(chart, f"K{chart_loc}")
        chart_loc += 15
    # Save the updated Excel file with the chart
    file_name = input("Input excel file name\n") + ".xlsx"
    wb.save(file_name)
    
def mag_to_excel(data_df):
    file_name = input("Input excel file name\n") + ".xlsx"
    wb = Workbook_xw(file_name)
    for ant_type in data_df['ant_type'].unique():
        df = data_df[data_df['ant_type'] == ant_type].sort_values(by='id').dropna(axis=1, how='all')
        ws = wb.add_worksheet(ant_type)
        for col_idx, col_name in enumerate(df.columns):
            ws.write_string(0, col_idx, str(col_name))
        for row_idx, row in enumerate(df.itertuples(index=False), start=1):
            for col_idx, val in enumerate(row):
                if isinstance(val, (int, float)):
                    if pd.isna(val) or np.isinf(val):
                        ws.write_blank(row_idx, col_idx, None)
                    else:
                        ws.write_number(row_idx, col_idx, val)
                elif val is not None:
                    ws.write_string(row_idx, col_idx, str(val))
    wb.close()

# Main execution
if __name__ == "__main__":
    try:
        user_input = choose('Processing phase or magnitude data?\n1 - Phase\n2 - Magnitude', '1', '2')
        if user_input == '1':
            process_type = 'phase'
        elif user_input == '2':
            process_type = 'magnitude'
            
        main_folder_path = summarize_folders()

        raw_df, failed_files = preprocess(main_folder_path, process_type)
        if failed_files:
            print("Failed to process the following files:")
            for file, error in failed_files:
                print(f"{file}: {error}")
        
        processed_df = process(raw_df, process_type)
        if process_type == 'phase':
            phase_to_excel(processed_df)
        elif process_type == 'magnitude':
            mag_to_excel(processed_df)
    except Exception as e:
        print(f"Error: {e}\nType: {type(e).__name__}\n")
        traceback.print_exc()
    input("Press Enter to exit\n")
